"""⑩ 成績一覧 Excel 出力（採点結果・各問判定・テスト分析・ランキング・成績一覧表）。"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from models.domain_repo import get_domain_column_labels, get_domain_max_score
from config import require_path_under_test_storage
from models.roster_repo import (
    get_roster_absent_state,
    get_roster_rows,
    get_selected_roster_name,
)
from models.test_repo import get_answer_fields, get_test_info
from services.feedback_renderer import _load_rows_with_extras
from services.grading import get_summary_data

ABSENT_LABEL = "未受験"
_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
_HEADER_FONT = Font(bold=True, size=10)
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11)


@dataclass
class StudentGradeRow:
    student_id: str
    name: str
    year: str = ""
    class_no: str = ""
    number: str = ""
    attr1: str = ""
    attr2: str = ""
    attr3: str = ""
    is_absent: bool = False
    external_score: float | None = None
    total_score: float | None = None
    hensachi: float | None = None
    domain_scores: dict[str, float] = field(default_factory=dict)
    judgments: dict[str, str] = field(default_factory=dict)
    scores: dict[str, Any] = field(default_factory=dict)


def _sid_sort_key(sid: str) -> tuple:
    s = str(sid or "").strip()
    if s.isdigit():
        return (0, int(s), s)
    return (1, 0, s.lower())


def _num_sort_key(value: str) -> tuple:
    s = str(value or "").strip()
    if s.isdigit():
        return (0, int(s))
    return (1, s)


def _has_any(rows: list[StudentGradeRow], attr: str) -> bool:
    return any(str(getattr(r, attr, "") or "").strip() for r in rows)


def _fmt_score(v: float | int | None) -> float | int:
    if v is None:
        return 0
    fv = float(v)
    if abs(fv - round(fv)) < 1e-9:
        return int(round(fv))
    return round(fv, 2)


def _parse_datetime_jp(raw: str) -> str:
    """実施日時文字列を『2026年7月3日実施』形式へ。"""
    text = str(raw or "").strip()
    if not text:
        return ""
    m = re.match(r"(\d{4})\D+(\d{1,2})\D+(\d{1,2})", text)
    if m:
        return f"{int(m.group(1))}年{int(m.group(2))}月{int(m.group(3))}日実施"
    return text


def format_ranking_title(subject: str, test_name: str, datetime_str: str) -> str:
    subj = str(subject or "").strip()
    name = str(test_name or "").strip()
    when = _parse_datetime_jp(datetime_str)
    head = "　".join(p for p in (subj, name) if p) or "テスト"
    if when:
        return f"{head}（{when}）"
    return head


def compute_hensachi(scores: list[float]) -> dict[float, float]:
    """得点値 → 偏差値。同一得点は同じ偏差値。σ=0 なら全員 50。"""
    if not scores:
        return {}
    n = len(scores)
    mean = sum(scores) / n
    var = sum((x - mean) ** 2 for x in scores) / n
    sd = math.sqrt(var)
    out: dict[float, float] = {}
    for x in scores:
        if sd < 1e-12:
            out[x] = 50.0
        else:
            out[x] = round(50.0 + 10.0 * (x - mean) / sd, 1)
    return out


def build_histogram_bins(
    max_score: int, bin_pct: int
) -> list[tuple[str, int, int]]:
    """(label, low_inclusive, high_inclusive) を満点側から降順。

    最初のビンは満点ちょうどだけ。以降は満点の bin_pct% 幅。
    """
    max_score = max(0, int(max_score))
    bin_pct = max(1, int(bin_pct))
    if max_score <= 0:
        return [("0", 0, 0)]
    step = max(1, int(round(max_score * bin_pct / 100.0)))
    bins: list[tuple[str, int, int]] = [(f"{max_score}", max_score, max_score)]
    high = max_score - 1
    while high >= 0:
        low = max(0, high - step + 1)
        if low == high:
            label = f"{low}"
        else:
            label = f"{high}-{low}"
        bins.append((label, low, high))
        high = low - 1
    return bins


def count_histogram(
    totals: list[float], bins: list[tuple[str, int, int]]
) -> list[tuple[str, int]]:
    counts = [0] * len(bins)
    for t in totals:
        score = int(round(float(t)))
        for i, (_lab, low, high) in enumerate(bins):
            if low <= score <= high:
                counts[i] += 1
                break
    return [(bins[i][0], counts[i]) for i in range(len(bins))]


def competition_rank(
    entries: list[tuple[float, str, str]],
) -> list[tuple[int, str, str, float]]:
    """(score, student_id, name) → (rank, id, name, score)。

    得点降順・同点は同一順位。同点内は ID 昇順表示。
    """
    ordered = sorted(entries, key=lambda e: (-float(e[0]), _sid_sort_key(e[1])))
    out: list[tuple[int, str, str, float]] = []
    prev_score: float | None = None
    rank = 0
    for i, (score, sid, name) in enumerate(ordered):
        if prev_score is None or abs(float(score) - float(prev_score)) > 1e-9:
            rank = i + 1
            prev_score = float(score)
        out.append((rank, sid, name, float(score)))
    return out


def build_student_grade_rows(test_id: str) -> tuple[list[StudentGradeRow], dict[str, Any]]:
    """名簿突合済みの生徒行とメタ情報。"""
    info = get_test_info(test_id)
    fields = get_answer_fields(test_id)
    points = info.get("points") or {}
    max_field_total = sum(int(points.get(f["id"], 0) or 0) for f in fields)
    domain_cols = get_domain_column_labels(test_id)

    result_rows = _load_rows_with_extras(test_id)
    by_sid: dict[str, dict[str, Any]] = {}
    for r in result_rows:
        sid = str(r.get("studentId") or "").strip()
        if sid:
            by_sid[sid] = r

    roster_name = get_selected_roster_name(test_id) or ""
    absent_state = get_roster_absent_state(test_id)
    absent_ids = {
        str(a.get("studentId") or "").strip()
        for a in (absent_state.get("absentStudents") or [])
        if a.get("studentId")
    }
    absent_names = {
        str(a.get("name") or "").strip()
        for a in (absent_state.get("absentStudents") or [])
        if a.get("name")
    }

    roster = get_roster_rows(roster_name) if roster_name else []
    students: list[StudentGradeRow] = []

    if roster:
        seen: set[str] = set()
        for rr in roster:
            sid = str(rr.get("studentId") or "").strip()
            name = str(rr.get("name") or "").strip()
            if not sid and not name:
                continue
            key = sid or f"name:{name}"
            if key in seen:
                continue
            seen.add(key)
            res = by_sid.get(sid) if sid else None
            marked_absent = (sid and sid in absent_ids) or (name and name in absent_names)
            is_absent = res is None or marked_absent
            row = StudentGradeRow(
                student_id=sid,
                name=name or str((res or {}).get("name") or ""),
                year=str(rr.get("year") or ""),
                class_no=str(rr.get("classNo") or ""),
                number=str(rr.get("number") or ""),
                attr1=str(rr.get("attr1") or ""),
                attr2=str(rr.get("attr2") or ""),
                attr3=str(rr.get("attr3") or ""),
                is_absent=is_absent,
            )
            if res is not None and not marked_absent:
                row.is_absent = False
                row.external_score = float(res.get("externalScore") or 0)
                row.total_score = float(res.get("totalScore") or 0)
                row.domain_scores = dict(res.get("domainScores") or {})
                row.judgments = dict(res.get("judgments") or {})
                row.scores = dict(res.get("scores") or {})
            students.append(row)
        # 名簿外だが results にある生徒も追加
        for sid, res in by_sid.items():
            if any(s.student_id == sid for s in students):
                continue
            students.append(
                StudentGradeRow(
                    student_id=sid,
                    name=str(res.get("name") or ""),
                    is_absent=False,
                    external_score=float(res.get("externalScore") or 0),
                    total_score=float(res.get("totalScore") or 0),
                    domain_scores=dict(res.get("domainScores") or {}),
                    judgments=dict(res.get("judgments") or {}),
                    scores=dict(res.get("scores") or {}),
                )
            )
    else:
        for res in result_rows:
            sid = str(res.get("studentId") or "").strip()
            students.append(
                StudentGradeRow(
                    student_id=sid,
                    name=str(res.get("name") or ""),
                    is_absent=False,
                    external_score=float(res.get("externalScore") or 0),
                    total_score=float(res.get("totalScore") or 0),
                    domain_scores=dict(res.get("domainScores") or {}),
                    judgments=dict(res.get("judgments") or {}),
                    scores=dict(res.get("scores") or {}),
                )
            )

    examinee_scores = [
        float(s.total_score)
        for s in students
        if not s.is_absent and s.total_score is not None
    ]
    hensachi_map = compute_hensachi(examinee_scores)
    for s in students:
        if not s.is_absent and s.total_score is not None:
            s.hensachi = hensachi_map.get(float(s.total_score))

    students.sort(key=lambda s: _sid_sort_key(s.student_id))

    max_total_observed = max(examinee_scores) if examinee_scores else 0.0
    perfect = max(int(max_field_total), int(math.ceil(max_total_observed)))

    meta = {
        "test_id": test_id,
        "test_name": str(info.get("testName") or ""),
        "subject": str(info.get("subject") or ""),
        "datetime": str(info.get("datetime") or ""),
        "fields": fields,
        "domain_cols": domain_cols,
        "max_score": perfect,
        "max_field_total": int(max_field_total),
        "title": format_ranking_title(
            str(info.get("subject") or ""),
            str(info.get("testName") or ""),
            str(info.get("datetime") or ""),
        ),
    }
    return students, meta


def _auto_width(ws, min_w: float = 4, max_w: float = 28) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, length + 2))


def _write_header(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_cell(ws, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _THIN
    if isinstance(value, (int, float)):
        cell.alignment = Alignment(horizontal="right")
    else:
        cell.alignment = Alignment(horizontal="center")


def _sheet_scores(wb: Workbook, students: list[StudentGradeRow], meta: dict) -> None:
    ws = wb.create_sheet("採点結果", 0)
    domain_cols: list[str] = list(meta.get("domain_cols") or [])
    show_year = _has_any(students, "year")
    show_class = _has_any(students, "class_no")
    show_number = _has_any(students, "number")
    show_a1 = _has_any(students, "attr1")
    show_a2 = _has_any(students, "attr2")
    show_a3 = _has_any(students, "attr3")
    show_ext = any(
        (not s.is_absent and s.external_score is not None and abs(float(s.external_score)) > 1e-9)
        for s in students
    )

    headers = ["4桁ID"]
    if show_year:
        headers.append("学年")
    if show_class:
        headers.append("組")
    if show_number:
        headers.append("番号")
    headers.append("氏名")
    if show_a1:
        headers.append("その他属性1")
    if show_a2:
        headers.append("その他属性2")
    if show_a3:
        headers.append("その他属性3")
    if show_ext:
        headers.append("外部得点")
    headers.extend(["総計点", "標準偏差値"])
    for col in domain_cols:
        headers.append(col.replace("_得点", "") if col.endswith("_得点") else col)

    _write_header(ws, 1, headers)
    for r_i, s in enumerate(students, 2):
        vals: list[Any] = [s.student_id]
        if show_year:
            vals.append(s.year)
        if show_class:
            vals.append(s.class_no)
        if show_number:
            vals.append(s.number)
        vals.append(s.name)
        if show_a1:
            vals.append(s.attr1)
        if show_a2:
            vals.append(s.attr2)
        if show_a3:
            vals.append(s.attr3)
        if s.is_absent:
            if show_ext:
                vals.append(ABSENT_LABEL)
            vals.append(ABSENT_LABEL)
            vals.append(ABSENT_LABEL)
            for _ in domain_cols:
                vals.append(ABSENT_LABEL)
        else:
            if show_ext:
                vals.append(_fmt_score(s.external_score))
            vals.append(_fmt_score(s.total_score))
            vals.append("" if s.hensachi is None else s.hensachi)
            for col in domain_cols:
                vals.append(_fmt_score(s.domain_scores.get(col, 0)))
        for c, v in enumerate(vals, 1):
            _write_cell(ws, r_i, c, v)
    _auto_width(ws)


def _sheet_judgments(wb: Workbook, students: list[StudentGradeRow], meta: dict) -> None:
    ws = wb.create_sheet("各問判定", 1)
    fields = list(meta.get("fields") or [])
    headers = ["4桁ID", "氏名"] + [
        str(f.get("displayName") or f.get("id") or "") for f in fields
    ]
    _write_header(ws, 1, headers)
    for r_i, s in enumerate(students, 2):
        vals: list[Any] = [s.student_id, s.name]
        for f in fields:
            fid = f["id"]
            if s.is_absent:
                vals.append(ABSENT_LABEL)
            else:
                j = str(s.judgments.get(fid) or "").strip()
                vals.append(j or "")
        for c, v in enumerate(vals, 1):
            _write_cell(ws, r_i, c, v)
    _auto_width(ws, max_w=18)


def _sheet_analysis(
    wb: Workbook,
    students: list[StudentGradeRow],
    meta: dict,
    *,
    hist_bin_pct: int,
) -> None:
    ws = wb.create_sheet("テスト分析", 2)
    summary = get_summary_data(str(meta["test_id"]))
    # なければ簡易再構築のため build_summary は呼ばない（重い／副作用）。summary 空なら表を現場計算。
    row = 1
    ws.cell(row=row, column=1, value="テスト分析").font = _TITLE_FONT
    row = 3
    ws.cell(row=row, column=1, value="各問題の正答率など").font = _SECTION_FONT
    row += 1
    q_headers = ["項目", "値", "備考"]
    _write_header(ws, row, q_headers)
    row += 1
    question_items = [s for s in summary if s.get("category") == "設問"]
    if not question_items:
        examinees = [s for s in students if not s.is_absent]
        n = len(examinees) or 1
        for f in meta.get("fields") or []:
            fid = f["id"]
            label = f.get("displayName") or fid
            o = sum(1 for s in examinees if str(s.judgments.get(fid) or "") == "○")
            question_items.append(
                {
                    "item": f"{label}_○率",
                    "value": f"{round(o / n * 1000) / 10}%",
                    "note": "",
                }
            )
    for item in question_items:
        _write_cell(ws, row, 1, item.get("item"))
        _write_cell(ws, row, 2, item.get("value"))
        _write_cell(ws, row, 3, item.get("note") or "")
        row += 1
    if not question_items:
        _write_cell(ws, row, 1, "（設問統計なし）")
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="大問・領域・範囲の得点率").font = _SECTION_FONT
    row += 1
    _write_header(ws, row, ["項目", "値", "備考"])
    row += 1
    domain_items = [s for s in summary if s.get("category") == "領域"]
    if not domain_items:
        for col in meta.get("domain_cols") or []:
            examinees = [s for s in students if not s.is_absent]
            vals = [float(s.domain_scores.get(col, 0) or 0) for s in examinees]
            max_sc = get_domain_max_score(str(meta["test_id"]), col)
            denom = (max_sc * (len(vals) or 1)) or 1
            rate = f"{round(sum(vals) / denom * 1000) / 10}%" if max_sc else "-"
            domain_items.append({"item": f"{col}_得点率", "value": rate, "note": f"満点={max_sc}"})
    for item in domain_items:
        _write_cell(ws, row, 1, item.get("item"))
        _write_cell(ws, row, 2, item.get("value"))
        _write_cell(ws, row, 3, item.get("note") or "")
        row += 1

    overall = [s for s in summary if s.get("category") == "全体"]
    if overall:
        row += 1
        ws.cell(row=row, column=1, value="全体").font = _SECTION_FONT
        row += 1
        _write_header(ws, row, ["項目", "値", "備考"])
        row += 1
        for item in overall:
            _write_cell(ws, row, 1, item.get("item"))
            _write_cell(ws, row, 2, item.get("value"))
            _write_cell(ws, row, 3, item.get("note") or "")
            row += 1

    # ヒストグラム
    row += 2
    ws.cell(row=row, column=1, value="総得点の分布（ヒストグラム）").font = _SECTION_FONT
    row += 1
    max_score = int(meta.get("max_score") or 0)
    bins = build_histogram_bins(max_score, hist_bin_pct)
    totals = [
        float(s.total_score)
        for s in students
        if not s.is_absent and s.total_score is not None
    ]
    hist = count_histogram(totals, bins)
    hist_header_row = row
    _write_header(ws, row, ["階級", "人数", "割合(%)"])
    row += 1
    hist_data_start = row
    total_n = len(totals) or 1
    for lab, cnt in hist:
        _write_cell(ws, row, 1, lab)
        _write_cell(ws, row, 2, cnt)
        _write_cell(ws, row, 3, round(cnt / total_n * 1000) / 10)
        row += 1
    hist_data_end = row - 1

    if hist_data_end >= hist_data_start:
        chart = BarChart()
        chart.type = "col"
        chart.title = "総得点分布"
        chart.y_axis.title = "人数"
        chart.x_axis.title = "階級"
        data = Reference(
            ws, min_col=2, min_row=hist_header_row, max_row=hist_data_end
        )
        cats = Reference(
            ws, min_col=1, min_row=hist_data_start, max_row=hist_data_end
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 18
        chart.height = 10
        ws.add_chart(chart, "E" + str(hist_header_row))

    _auto_width(ws, max_w=40)


def _sheet_ranking(
    wb: Workbook,
    students: list[StudentGradeRow],
    meta: dict,
    *,
    overall_limit: int,
    class_limit: int,
) -> None:
    ws = wb.create_sheet("ランキング", 3)
    title = str(meta.get("title") or "")
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    examinees = [s for s in students if not s.is_absent and s.total_score is not None]
    entries = [
        (float(s.total_score), s.student_id, s.name) for s in examinees
    ]
    ranked = competition_rank(entries)

    row = 3
    ws.cell(row=row, column=1, value=f"全体（上位{overall_limit}位）").font = _SECTION_FONT
    row += 1
    _write_header(ws, row, ["順位", "4桁ID", "氏名", "総計点"])
    row += 1
    for rank, sid, name, score in ranked:
        if rank > overall_limit:
            break
        _write_cell(ws, row, 1, rank)
        _write_cell(ws, row, 2, sid)
        _write_cell(ws, row, 3, name)
        _write_cell(ws, row, 4, _fmt_score(score))
        row += 1

    # クラス別
    classes = sorted(
        {
            str(s.class_no or "").strip()
            for s in examinees
            if str(s.class_no or "").strip()
        },
        key=_num_sort_key,
    )
    for class_no in classes:
        row += 2
        ws.cell(
            row=row, column=1, value=f"{class_no}組（上位{class_limit}位）"
        ).font = _SECTION_FONT
        row += 1
        _write_header(ws, row, ["順位", "4桁ID", "氏名", "総計点"])
        row += 1
        class_entries = [
            (float(s.total_score), s.student_id, s.name)
            for s in examinees
            if str(s.class_no or "").strip() == class_no
        ]
        for rank, sid, name, score in competition_rank(class_entries):
            if rank > class_limit:
                break
            _write_cell(ws, row, 1, rank)
            _write_cell(ws, row, 2, sid)
            _write_cell(ws, row, 3, name)
            _write_cell(ws, row, 4, _fmt_score(score))
            row += 1

    _auto_width(ws)


def _sheet_class_list(wb: Workbook, students: list[StudentGradeRow], meta: dict) -> None:
    ws = wb.create_sheet("成績一覧表", 4)
    title = str(meta.get("title") or "")
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT

    # クラス順（空組は最後）
    by_class: dict[str, list[StudentGradeRow]] = {}
    for s in students:
        key = str(s.class_no or "").strip() or "（組なし）"
        by_class.setdefault(key, []).append(s)
    class_keys = sorted(
        by_class.keys(),
        key=lambda k: (1, k) if k == "（組なし）" else (0, _num_sort_key(k)),
    )[:6]

    for key in class_keys:
        by_class[key].sort(
            key=lambda s: (_num_sort_key(s.number), _sid_sort_key(s.student_id))
        )

    # 各クラスを横並びに 3 列（ID・氏名・総計）
    cols_per_class = 3
    gap = 1

    for i, key in enumerate(class_keys):
        base = 1 + i * (cols_per_class + gap)
        header = f"{key}組" if key != "（組なし）" else key
        cell = ws.cell(row=3, column=base, value=header)
        cell.font = _SECTION_FONT
        ws.merge_cells(
            start_row=3, start_column=base, end_row=3, end_column=base + 2
        )
        for offset, htxt in enumerate(["4桁ID", "氏名", "総計点"]):
            c = ws.cell(row=4, column=base + offset, value=htxt)
            c.font = _HEADER_FONT
            c.border = _THIN
            c.alignment = Alignment(horizontal="center")

        for r_i, s in enumerate(by_class[key]):
            rr = 5 + r_i
            total_v: Any = ABSENT_LABEL if s.is_absent else _fmt_score(s.total_score)
            for offset, val in enumerate([s.student_id, s.name, total_v]):
                _write_cell(ws, rr, base + offset, val)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5, header=0.2, footer=0.2
    )
    ws.print_title_rows = "1:4"

    # 列幅コンパクト
    for i, _key in enumerate(class_keys):
        base = 1 + i * (cols_per_class + gap)
        ws.column_dimensions[get_column_letter(base)].width = 8
        ws.column_dimensions[get_column_letter(base + 1)].width = 10
        ws.column_dimensions[get_column_letter(base + 2)].width = 7
        if i < len(class_keys) - 1:
            ws.column_dimensions[get_column_letter(base + 3)].width = 2


def export_grade_excel(
    test_id: str,
    output_path: str | Path,
    *,
    prefs: dict | None = None,
) -> Path:
    """成績 Excel を生成して保存する。"""
    prefs = prefs or load_excel_export_prefs()
    students, meta = build_student_grade_rows(test_id)
    if not students:
        raise ValueError("出力する生徒データがありません。名簿または採点結果を確認してください。")

    wb = Workbook()
    # デフォルトシートを除去
    default = wb.active
    wb.remove(default)

    _sheet_scores(wb, students, meta)
    _sheet_judgments(wb, students, meta)
    _sheet_analysis(
        wb, students, meta, hist_bin_pct=int(prefs["hist_bin_pct"])
    )
    _sheet_ranking(
        wb,
        students,
        meta,
        overall_limit=int(prefs["rank_overall_limit"]),
        class_limit=int(prefs["rank_class_limit"]),
    )
    _sheet_class_list(wb, students, meta)

    path = require_path_under_test_storage(test_id, output_path, label="Excel の保存先")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
