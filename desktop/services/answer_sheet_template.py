"""回答用紙ひな形 Excel（GAS ハブSS のテンプレート_共通A4横/縦 と同等）。"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

SHEET_TEMPLATE_A4_LANDSCAPE = "テンプレート_共通A4横"
SHEET_TEMPLATE_A4_PORTRAIT = "テンプレート_共通A4縦"

# GAS code.gs と同じオフセット・グリッド定義（1マス ≈ 20px）
TEMPLATE_GRID_OFFSET_ROW = 2  # 0-based offset → start row = 3
TEMPLATE_GRID_OFFSET_COL = 2

_THIN = Side(style="thin", color="CCCCCC")
_THICK = Side(style="medium", color="000000")
_MARK_THIN = Side(style="thin", color="AAAAAA")
_MARK_OUTER = Side(style="thin", color="64748B")
_BLACK = Side(style="thin", color="000000")


def get_paper_template_config(orientation: str) -> dict[str, Any]:
    is_portrait = orientation == "portrait"
    return {
        "orientation": orientation,
        "borderCols": 51 if is_portrait else 73,
        "borderRows": 73 if is_portrait else 51,
        "idStartCol": 42 if is_portrait else 64,
        "idStartRow": 5,
    }


def _apply_grid(ws, cfg: dict[str, Any]) -> dict[str, int]:
    start_row = TEMPLATE_GRID_OFFSET_ROW + 1
    start_col = TEMPLATE_GRID_OFFSET_COL + 1
    end_row = start_row + int(cfg["borderRows"]) - 1
    end_col = start_col + int(cfg["borderCols"]) - 1

    # GAS: マス目列幅 21px / 行高 15px、左2列は少し広め
    for c in range(1, end_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 2.6 if c >= start_col else 4.5
    for r in range(1, end_row + 1):
        ws.row_dimensions[r].height = 11.25

    white = PatternFill("solid", fgColor="FFFFFF")
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c, "")
            cell.fill = white
            left = _THICK if c == start_col else _THIN
            right = _THICK if c == end_col else _THIN
            top = _THICK if r == start_row else _THIN
            bottom = _THICK if r == end_row else _THIN
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    return {
        "startRow": start_row,
        "startCol": start_col,
        "endRow": end_row,
        "endCol": end_col,
    }


def _build_student_id_mark_block(ws, cfg: dict[str, Any], id_digits: int = 4) -> None:
    """参考マークシートと同じ生徒IDマーク欄（年/組/番 + 0〜9）。"""
    id_start_col = int(cfg["idStartCol"])
    id_start_row = int(cfg["idStartRow"])
    label_col = id_start_col - 1
    header_col = id_start_col - 4
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    gray_fill = PatternFill("solid", fgColor="F3F3F3")
    black_border = Border(left=_BLACK, right=_BLACK, top=_BLACK, bottom=_BLACK)
    mark_border = Border(left=_MARK_THIN, right=_MARK_THIN, top=_MARK_THIN, bottom=_MARK_THIN)

    # 生徒ID 見出し（3列×idDigits行を結合）
    ws.merge_cells(
        start_row=id_start_row,
        start_column=header_col,
        end_row=id_start_row + id_digits - 1,
        end_column=header_col + 2,
    )
    header = ws.cell(id_start_row, header_col, "生徒\nID")
    header.font = Font(bold=True, size=9)
    header.fill = gray_fill
    header.alignment = center
    for r in range(id_start_row, id_start_row + id_digits):
        for c in range(header_col, header_col + 3):
            ws.cell(r, c).border = black_border
            ws.cell(r, c).fill = gray_fill

    # 年・組・番
    year = ws.cell(id_start_row, label_col, "年")
    year.font = Font(size=8)
    year.alignment = center
    year.border = black_border
    klass = ws.cell(id_start_row + 1, label_col, "組")
    klass.font = Font(size=8)
    klass.alignment = center
    klass.border = black_border
    if id_digits > 2:
        ws.merge_cells(
            start_row=id_start_row + 2,
            start_column=label_col,
            end_row=id_start_row + id_digits - 1,
            end_column=label_col,
        )
        ban = ws.cell(id_start_row + 2, label_col, "番")
        ban.font = Font(size=8)
        ban.alignment = center
        for r in range(id_start_row + 2, id_start_row + id_digits):
            ws.cell(r, label_col).border = black_border
            ws.cell(r, label_col).alignment = center

    # 0〜9 塗りつぶし欄
    for r in range(id_digits):
        for c in range(10):
            cell = ws.cell(id_start_row + r, id_start_col + c, str(c))
            cell.alignment = center
            cell.font = Font(color="AAAAAA", size=9)
            cell.border = mark_border

    # ブロック外枠
    block_end_row = id_start_row + id_digits - 1
    block_end_col = id_start_col + 9
    for r in range(id_start_row, block_end_row + 1):
        for c in range(header_col, block_end_col + 1):
            cell = ws.cell(r, c)
            b = cell.border
            cell.border = Border(
                left=_MARK_OUTER if c == header_col else (b.left or _MARK_THIN),
                right=_MARK_OUTER if c == block_end_col else (b.right or _MARK_THIN),
                top=_MARK_OUTER if r == id_start_row else (b.top or _MARK_THIN),
                bottom=_MARK_OUTER if r == block_end_row else (b.bottom or _MARK_THIN),
            )


def _build_sheet(ws, orientation: str) -> None:
    cfg = get_paper_template_config(orientation)
    grid = _apply_grid(ws, cfg)
    orient_label = "A4縦" if orientation == "portrait" else "A4横"

    title_end = min(grid["endCol"], 40)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=title_end)
    title = ws.cell(
        1,
        1,
        f"【回答用紙ひな形 {orient_label}】このシートを編集して印刷してください。"
        "外枠・マス目・IDマーク欄の位置は ① 回答欄設定の座標系（1マス≈20px）と一致しています。",
    )
    title.font = Font(size=9)
    title.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 28

    note = ws.cell(
        2,
        grid["startCol"],
        "→ 記述欄を追加する場合はマス目内を編集。印刷後スキャンして ③ テキスト化で読み込みます。",
    )
    note.font = Font(size=8, color="64748B")

    _build_student_id_mark_block(ws, cfg, 4)
    ws.freeze_panes = ws.cell(grid["startRow"], 1)

    # 印刷設定（A4・向き）
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait" if orientation == "portrait" else "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.4, bottom=0.4)
    ws.print_area = f"A1:{get_column_letter(grid['endCol'])}{grid['endRow']}"


def export_answer_sheet_templates(output_path: str) -> str:
    """A4横・A4縦の回答用紙ひな形（生徒IDマーク欄付き）を Excel に出力。

    GAS の「テンプレート_共通A4横」「テンプレート_共通A4縦」とほぼ同一の書式。
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # 既定シートを横向きに
    ws_land = wb.active
    ws_land.title = SHEET_TEMPLATE_A4_LANDSCAPE
    _build_sheet(ws_land, "landscape")

    ws_port = wb.create_sheet(SHEET_TEMPLATE_A4_PORTRAIT)
    _build_sheet(ws_port, "portrait")

    wb.save(path)
    return str(path.resolve())
